from bs4 import BeautifulSoup
import re
import os
import os.path
import urllib.request
import urllib.parse
import csv
from  openpyxl import * 

filename=[]
rootdir='data'
for parent,dirnames,filenames in os.walk(rootdir):
    for name in filenames:
        filename.append(str(name))
        print(name)
    for name in filename:
        print(name)
wb=Workbook()

ws=wb.active

ws.title='cnames'
ws['A1']='Source'
ws['B1']='Target'
n=0
m=0
l=len(filename)
for s in filename[1:]:
    with open('data/%s'%(str(s)),'r') as f:
        print(s)        
        word=BeautifulSoup(f.read(),'lxml')
    word1=word.find_all('a',attrs={'class':'in-block vertival-middle overflow-width'})
    word2=BeautifulSoup(str(word1),'lxml')
    r=re.compile('CompangyDetail\.gudong\..+?')
    word3=word2.find_all('a',attrs={'tyc-event-ch':r})
    r=re.compile('<title>(.+?)_')
    word_cname=(re.search(r,str(word))).group(1)
    print(word_cname)
    print(word3)
    title=[]
    title_gudong=[]
    i=0
    for content in word3:
        r=re.compile('title="(.+?)"')
        title_gudong.append(re.search(r,str(content)))
            
        if title_gudong[i].group(1)!=None:
            title_gudong[i]=str(title_gudong[i].group(1))
        #else:
           # title_gudong[i]='空'
        ws['A%s'%(m+2)]=word_cname
        ws['B%s'%(m+2)]=title_gudong[i]
        i=i+1
        m=m+1
        print(title_gudong[i-1])         
    title.append({'name':str(word_cname),'gudong':title_gudong})
            
            
    #print(title[n]['name'])
        
    #ws['A1']='Label'
    #ws['A%s'%(n+1)]=word_cname
    print('当前进度：%f%%'%(n/l*100))       
    n=n+1
wb.save('testbian.cvs')
